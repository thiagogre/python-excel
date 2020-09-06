import pandas as pd
from openpyxl import load_workbook
from collections import Counter
import numpy as np

excel_file = 'C:/Users/thiag/Desktop/Atlas/QMtest.xlsx'
excel_file_inline = 'C:/Users/thiag/Desktop/Atlas/inline.xlsx'
loading = load_workbook(filename=excel_file_inline)
la = loading.active
la.delete_cols(7)
la.delete_cols(7)
la.delete_cols(7)
la.delete_cols(7)
la.delete_cols(8)
la.delete_cols(8)
la.delete_cols(8)
la.delete_cols(8)
la.delete_cols(5)
loading.save(excel_file_inline)

dfInline = pd.read_excel(excel_file_inline)
loading_1 = load_workbook(filename=excel_file_inline)
la_1 = loading_1.active
compensating_value = 2
repeated_notes = {}
copy = ''
aux = 0

for note in dfInline['Nota']:

    repeated_notes[note] = Counter(dfInline['Nota'])[note]

    if (repeated_notes[note] == 2) and (copy != note):
        measure_text = 'E'+str(compensating_value)
        la_1.move_range(measure_text, rows=1, cols=2)
        measure_code_text = 'F'+str(compensating_value)
        la_1.move_range(measure_code_text, rows=1, cols=2)
        copy = note

    if (repeated_notes[note] == 3) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=3)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=3)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux = 0

    if (repeated_notes[note] == 4) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=2)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=4)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux += 1
                if aux == 2:
                    measure_text = 'E'+str(compensating_value+3)
                    la_1.move_range(measure_text, rows=-3, cols=6)
                    measure_code_text = 'F'+str(compensating_value+3)
                    la_1.move_range(measure_code_text, rows=-3, cols=6)
                    aux = 0

    if (repeated_notes[note] == 5) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=2)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=4)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux += 1
                if aux == 2:
                    measure_text = 'E'+str(compensating_value+3)
                    la_1.move_range(measure_text, rows=-3, cols=6)
                    measure_code_text = 'F'+str(compensating_value+3)
                    la_1.move_range(measure_code_text, rows=-3, cols=6)
                    aux += 1
                    if aux == 3:
                        measure_text = 'E'+str(compensating_value+4)
                        la_1.move_range(measure_text, rows=-4, cols=8)
                        measure_code_text = 'F'+str(compensating_value+4)
                        la_1.move_range(measure_code_text, rows=-4, cols=8)
                        aux = 0
    if (repeated_notes[note] == 6) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=2)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=4)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux += 1
                if aux == 2:
                    measure_text = 'E'+str(compensating_value+3)
                    la_1.move_range(measure_text, rows=-3, cols=6)
                    measure_code_text = 'F'+str(compensating_value+3)
                    la_1.move_range(measure_code_text, rows=-3, cols=6)
                    aux += 1
                    if aux == 3:
                        measure_text = 'E'+str(compensating_value+4)
                        la_1.move_range(measure_text, rows=-4, cols=8)
                        measure_code_text = 'F'+str(compensating_value+4)
                        la_1.move_range(measure_code_text, rows=-4, cols=8)
                        aux += 1
                        if aux == 4:
                            measure_text = 'E'+str(compensating_value+5)
                            la_1.move_range(measure_text, rows=-5, cols=10)
                            measure_code_text = 'F'+str(compensating_value+5)
                            la_1.move_range(
                                measure_code_text, rows=-5, cols=10)
                            aux = 0
    if (repeated_notes[note] == 7) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=2)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=4)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux += 1
                if aux == 2:
                    measure_text = 'E'+str(compensating_value+3)
                    la_1.move_range(measure_text, rows=-3, cols=6)
                    measure_code_text = 'F'+str(compensating_value+3)
                    la_1.move_range(measure_code_text, rows=-3, cols=6)
                    aux += 1
                    if aux == 3:
                        measure_text = 'E'+str(compensating_value+4)
                        la_1.move_range(measure_text, rows=-4, cols=8)
                        measure_code_text = 'F'+str(compensating_value+4)
                        la_1.move_range(measure_code_text, rows=-4, cols=8)
                        aux += 1
                        if aux == 4:
                            measure_text = 'E'+str(compensating_value+5)
                            la_1.move_range(measure_text, rows=-5, cols=10)
                            measure_code_text = 'F'+str(compensating_value+5)
                            la_1.move_range(
                                measure_code_text, rows=-5, cols=10)
                            aux += 1
                            if aux == 5:
                                measure_text = 'E'+str(compensating_value+6)
                                la_1.move_range(measure_text, rows=-6, cols=12)
                                measure_code_text = 'F'+str(
                                    compensating_value+6)
                                la_1.move_range(
                                    measure_code_text, rows=-6, cols=12)
                                aux = 0
    if (repeated_notes[note] == 8) and (copy != note):
        if aux == 0:
            measure_text = 'E'+str(compensating_value+1)
            la_1.move_range(measure_text, rows=-1, cols=2)
            measure_code_text = 'F'+str(compensating_value+1)
            la_1.move_range(measure_code_text, rows=-1, cols=2)
            aux += 1
            copy = note
            if aux == 1:
                measure_text = 'E'+str(compensating_value+2)
                la_1.move_range(measure_text, rows=-2, cols=4)
                measure_code_text = 'F'+str(compensating_value+2)
                la_1.move_range(measure_code_text, rows=-2, cols=4)
                aux += 1
                if aux == 2:
                    measure_text = 'E'+str(compensating_value+3)
                    la_1.move_range(measure_text, rows=-3, cols=6)
                    measure_code_text = 'F'+str(compensating_value+3)
                    la_1.move_range(measure_code_text, rows=-3, cols=6)
                    aux += 1
                    if aux == 3:
                        measure_text = 'E'+str(compensating_value+4)
                        la_1.move_range(measure_text, rows=-4, cols=8)
                        measure_code_text = 'F'+str(compensating_value+4)
                        la_1.move_range(measure_code_text, rows=-4, cols=8)
                        aux += 1
                        if aux == 4:
                            measure_text = 'E'+str(compensating_value+5)
                            la_1.move_range(measure_text, rows=-5, cols=10)
                            measure_code_text = 'F'+str(compensating_value+5)
                            la_1.move_range(
                                measure_code_text, rows=-5, cols=10)
                            aux += 1
                            if aux == 5:
                                measure_text = 'E'+str(compensating_value+6)
                                la_1.move_range(measure_text, rows=-6, cols=12)
                                measure_code_text = 'F'+str(
                                    compensating_value+6)
                                la_1.move_range(
                                    measure_code_text, rows=-6, cols=12)
                                aux += 1
                                if aux == 6:
                                    measure_text = 'E'+str(
                                        compensating_value+7)
                                    la_1.move_range(
                                        measure_text, rows=-7, cols=14)
                                    measure_code_text = 'F'+str(
                                        compensating_value+7)
                                    la_1.move_range(
                                        measure_code_text, rows=-7, cols=14)
                                    aux = 0
    compensating_value += 1

loading_1.save(excel_file_inline)
df = pd.read_excel(excel_file_inline)
df['Texto das medidas'].replace('', np.nan, inplace=True)
df.dropna(subset=['Texto das medidas'], inplace=True)
df.rename(columns={
    'Unnamed: 6': 'Tipo NF',
    'Unnamed: 7': 'Emissão NF',
    'Unnamed: 8': 'Retorno NF',
    'Unnamed: 9': 'Conhecimento',
    'Unnamed: 10': 'Valor',
    'Unnamed: 11': 'Conhecimento',
    'Unnamed: 12': 'Relatório',
    'Unnamed: 13': 'Conhecimento',
    'Unnamed: 14': 'Material',
    'Unnamed: 15': 'Avaliar',
    'Unnamed: 16': '',
    'Unnamed: 17': '',
    'Unnamed: 18': '',
    'Unnamed: 19': '',
}, inplace=True)

df.to_excel(excel_file_inline)
